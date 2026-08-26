# -*- coding: utf-8 -*-
# pip install openpyxl pillow
"""
Office document helper toolkit.

Sub-commands:
- word function   : Placeholder namespace for Word tools.
- excel function  : Excel tools, including worksheet image export.
- slides function : Placeholder namespace for slide tools.
"""

import argparse
import colorsys
import inspect
import re
import sys
import traceback
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Protocol
from xml.etree import ElementTree

try:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.styles.colors import COLOR_INDEX
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import range_boundaries
    from PIL import Image, ImageDraw, ImageFont
except ImportError as exc:
    load_workbook = None
    MergedCell = None
    COLOR_INDEX = ()
    get_column_letter = None
    range_boundaries = None
    Image = None
    ImageDraw = None
    ImageFont = None
    OFFICE_IMPORT_ERROR: ImportError | None = exc
else:
    OFFICE_IMPORT_ERROR = None


DEBUG_MODE = False

DPI_SCALE = 96 / 72
CELL_PADDING_X = 8
CELL_PADDING_Y = 5
MIN_COLUMN_WIDTH = 36
MIN_RENDER_ROW_HEIGHT = 1
WIDTH_FACTOR = 1.35
LINE_SPACING = 5
DEFAULT_BACKGROUND = "#ffffff"
DEFAULT_GRID_COLOR = "#b7b7b7"
DEFAULT_TEXT_COLOR = "#1f1f1f"
DEFAULT_BAND_COLORS = ("#ffffff", "#f5f9fc")
DEFAULT_THEME_HEADER_TEXT = "\u4e3b\u9898"
DEFAULT_FONT_SIZE = 11
DEFAULT_ROW_HEIGHT_POINTS = 15.0
DEFAULT_COLUMN_WIDTH = 8.43
FONT_FAMILY_PATHS = {
    "\u5b8b\u4f53": (
        "/home/wkyuu/.local/share/fonts/STSONG.TTF",
        "/home/wkyuu/.local/share/fonts/msyh.ttc",
    ),
    "SimSun": ("/home/wkyuu/.local/share/fonts/STSONG.TTF",),
    "STSong": ("/home/wkyuu/.local/share/fonts/STSONG.TTF",),
    "\u534e\u6587\u5b8b\u4f53": ("/home/wkyuu/.local/share/fonts/STSONG.TTF",),
    "Carlito": ("/home/wkyuu/.local/share/fonts/calibri.ttf",),
    "Calibri": ("/home/wkyuu/.local/share/fonts/calibri.ttf",),
}
FONT_FAMILY_BOLD_PATHS = {
    "\u5b8b\u4f53": (
        "/home/wkyuu/.local/share/fonts/msyhbd.ttc",
        "/home/wkyuu/.local/share/fonts/STSONG.TTF",
    ),
    "SimSun": (
        "/home/wkyuu/.local/share/fonts/msyhbd.ttc",
        "/home/wkyuu/.local/share/fonts/STSONG.TTF",
    ),
    "STSong": (
        "/home/wkyuu/.local/share/fonts/msyhbd.ttc",
        "/home/wkyuu/.local/share/fonts/STSONG.TTF",
    ),
    "\u534e\u6587\u5b8b\u4f53": (
        "/home/wkyuu/.local/share/fonts/msyhbd.ttc",
        "/home/wkyuu/.local/share/fonts/STSONG.TTF",
    ),
    "Carlito": ("/home/wkyuu/.local/share/fonts/calibrib.ttf",),
    "Calibri": ("/home/wkyuu/.local/share/fonts/calibrib.ttf",),
}
DEFAULT_FONT_PATHS = (
    "/home/wkyuu/.local/share/fonts/msyh.ttc",
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
)
DEFAULT_BOLD_FONT_PATHS = (
    "/home/wkyuu/.local/share/fonts/msyhbd.ttc",
    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
)


class CLIStyle:
    """CLI tool unified style config."""

    COLORS = {
        "TITLE": 7,
        "SUB_TITLE": 2,
        "CONTENT": 3,
        "EXAMPLE": 7,
        "WARNING": 4,
        "ERROR": 2,
        "INFO": 5,
        "OK": 3,
    }

    @staticmethod
    def color(text: str = "", color: int = COLORS["CONTENT"]) -> str:
        """Apply terminal color to text."""
        color_table = {
            0: "{}",
            1: "\033[1;30m{}\033[0m",
            2: "\033[1;31m{}\033[0m",
            3: "\033[1;32m{}\033[0m",
            4: "\033[1;33m{}\033[0m",
            5: "\033[1;34m{}\033[0m",
            6: "\033[1;35m{}\033[0m",
            7: "\033[1;36m{}\033[0m",
            8: "\033[1;37m{}\033[0m",
        }
        return color_table[color].format(text)

    @staticmethod
    def write(
        text: str = "", color: int = COLORS["CONTENT"], error: bool = False
    ) -> None:
        """Write a styled line to stdout or stderr."""
        stream = sys.stderr if error else sys.stdout
        stream.write(f"{CLIStyle.color(text, color)}\n")


class ColoredArgumentParser(argparse.ArgumentParser):
    """Argument parser with colored help output."""

    def _format_action_invocation(self, action: argparse.Action) -> str:
        """Format option names with semantic terminal colors."""
        if not action.option_strings:
            (metavar,) = self._metavar_formatter(action, action.dest)(1)
            return metavar

        parts = []
        if action.nargs == 0:
            parts.extend(
                CLIStyle.color(option, CLIStyle.COLORS["SUB_TITLE"])
                for option in action.option_strings
            )
            return ", ".join(parts)

        args_string = self._format_args(action, action.dest.upper())
        for option_string in action.option_strings:
            parts.append(
                CLIStyle.color(
                    f"{option_string} {args_string}",
                    CLIStyle.COLORS["SUB_TITLE"],
                )
            )
        return ", ".join(parts)

    def format_help(self) -> str:
        """Return help text with colored descriptions and sections."""
        formatter = self._get_formatter()
        if self.description:
            formatter.add_text(
                CLIStyle.color(self.description, CLIStyle.COLORS["TITLE"])
            )
        formatter.add_usage(self.usage, self._actions, self._mutually_exclusive_groups)
        formatter.add_text(
            CLIStyle.color("\nOptional Arguments:", CLIStyle.COLORS["TITLE"])
        )
        for action_group in self._action_groups:
            formatter.start_section(action_group.title)
            formatter.add_arguments(action_group._group_actions)
            formatter.end_section()
        if self.epilog:
            formatter.add_text(self.epilog)
        return formatter.format_help()


@dataclass(frozen=True)
class ImageRenderOptions:
    """Configuration for raster image output."""

    scale: int
    background: str
    grid_color: str
    text_color: str
    theme_header_text: str
    use_theme_bands: bool
    preserve_fills: bool
    theme_colors: dict[int, str] = field(default_factory=dict)


@dataclass(frozen=True)
class ImageConversionRequest:
    """Input and output paths for an image conversion command."""

    input_path: Path
    output_dir: Path
    scale: int
    selected_sheets: tuple[str, ...]
    use_theme_bands: bool
    preserve_fills: bool
    theme_header_text: str


@dataclass(frozen=True)
class ImageConversionResult:
    """Result reported by a document-to-image conversion."""

    output_paths: tuple[Path, ...]


class ImageConverter(Protocol):
    """Common interface for document image converters."""

    def convert(self, request: ImageConversionRequest) -> ImageConversionResult:
        """Convert one document to image files."""


def debug(*args: Any, **kwargs: Any) -> None:
    """
    Print debug information with source file and line number.
    ```python
    debug("value", object_value)

    return = None
    ```
    """
    if not DEBUG_MODE:
        return
    frame = inspect.currentframe()
    caller = frame.f_back if frame else None
    if caller is None:
        CLIStyle.write("[DEBUG]", CLIStyle.COLORS["INFO"], error=True)
        return
    prefix = f"[DEBUG {Path(caller.f_code.co_filename).name}:{caller.f_lineno}]"
    message = " ".join(str(arg) for arg in args)
    CLIStyle.write(f"{prefix} {message}", CLIStyle.COLORS["INFO"], error=True)
    if kwargs:
        CLIStyle.write(str(kwargs), CLIStyle.COLORS["INFO"], error=True)


def create_example_text(
    script_name: str,
    examples: list[tuple[str, str]],
    notes: list[str] | None = None,
) -> str:
    """Create colored example and note text for argparse epilog blocks."""
    text = f"\n{CLIStyle.color('Examples:', CLIStyle.COLORS['SUB_TITLE'])}"
    for desc, cmd in examples:
        text += f"\n  {CLIStyle.color(f'# {desc}', CLIStyle.COLORS['EXAMPLE'])}"
        text += f"\n  {CLIStyle.color(f'{script_name} {cmd}', CLIStyle.COLORS['CONTENT'])}\n"
    if notes:
        text += f"\n{CLIStyle.color('Notes:', CLIStyle.COLORS['SUB_TITLE'])}"
        for note in notes:
            text += f"\n  {CLIStyle.color(f'- {note}', CLIStyle.COLORS['CONTENT'])}"
    return text


def ensure_office_dependencies() -> None:
    """Raise an actionable error when optional Office dependencies are missing."""
    if OFFICE_IMPORT_ERROR is not None:
        raise RuntimeError(
            "Missing dependency. Install with: pip install openpyxl pillow"
        ) from OFFICE_IMPORT_ERROR


def find_font_path(font_name: str | None = None, bold: bool = False) -> str | None:
    """Return the first available path for a worksheet font."""
    family_paths = FONT_FAMILY_BOLD_PATHS if bold else FONT_FAMILY_PATHS
    fallback_paths = DEFAULT_BOLD_FONT_PATHS if bold else DEFAULT_FONT_PATHS
    font_paths = family_paths.get(font_name or "", ()) + fallback_paths
    for font_path in font_paths:
        if Path(font_path).exists():
            return font_path
    return None


def load_render_font(
    size: int, font_name: str | None = None, bold: bool = False
) -> Any:
    """Load a font for worksheet rendering."""
    ensure_office_dependencies()
    font_path = find_font_path(font_name, bold=bold)
    if font_path is None:
        return ImageFont.load_default()
    return ImageFont.truetype(font_path, size=size, index=0)


def sanitize_filename(value: str) -> str:
    """Make a document object name safe as a file name."""
    safe_name = re.sub(r'[\\/:*?"<>|]+', "_", value).strip()
    safe_name = re.sub(r"\s+", "_", safe_name)
    return safe_name or "document"


def points_to_pixels(
    points: float | None, default_points: float = DEFAULT_ROW_HEIGHT_POINTS
) -> int:
    """Convert point height to image pixels."""
    effective_points = default_points if points is None else points
    return max(1, int(round(effective_points * DPI_SCALE)))


def column_width_to_pixels(
    width: float | None, default_width: float = DEFAULT_COLUMN_WIDTH
) -> int:
    """Convert spreadsheet column width to approximate image pixels."""
    effective_width = default_width if width is None else width
    return max(
        MIN_COLUMN_WIDTH,
        int(round((effective_width * 7 + 8) * WIDTH_FACTOR)),
    )


def parse_theme_colors(theme_xml: bytes | None) -> dict[int, str]:
    """Parse theme color scheme entries into indexed RGB colors."""
    if not theme_xml:
        return {}

    try:
        root = ElementTree.fromstring(theme_xml)
    except (ElementTree.ParseError, TypeError):
        return {}

    color_scheme = next(
        (
            element
            for element in root.iter()
            if element.tag.rsplit("}", 1)[-1] == "clrScheme"
        ),
        None,
    )
    if color_scheme is None:
        return {}

    theme_colors = {}
    for index, scheme_color in enumerate(color_scheme):
        color_node = next(iter(scheme_color), None)
        if color_node is None:
            continue

        color_type = color_node.tag.rsplit("}", 1)[-1]
        raw_color = color_node.attrib.get("val")
        if color_type == "sysClr":
            raw_color = color_node.attrib.get("lastClr", raw_color)
        if raw_color is None or not re.fullmatch(r"[0-9a-fA-F]{6,8}", raw_color):
            continue
        theme_colors[index] = f"#{raw_color[-6:]}".upper()

    return theme_colors


def apply_excel_tint(rgb: str, tint: float) -> str:
    """Apply Excel's HLS tint transformation to an RGB color."""
    red = int(rgb[1:3], 16) / 255
    green = int(rgb[3:5], 16) / 255
    blue = int(rgb[5:7], 16) / 255
    hue, lightness, saturation = colorsys.rgb_to_hls(red, green, blue)

    if tint < 0:
        lightness *= 1 + tint
    else:
        lightness = lightness * (1 - tint) + tint

    lightness = max(0.0, min(1.0, lightness))
    red, green, blue = colorsys.hls_to_rgb(hue, lightness, saturation)
    return "#%02X%02X%02X" % (
        round(red * 255),
        round(green * 255),
        round(blue * 255),
    )


def get_color(value: Any, theme_colors: dict[int, str] | None = None) -> str | None:
    """Convert an openpyxl color value, including theme colors, to hex."""
    if value is None or value.type is None:
        return None

    if value.type == "rgb" and value.rgb:
        rgb = value.rgb[-6:]
        return f"#{rgb}".upper()

    if value.type == "indexed" and value.indexed is not None:
        index = int(value.indexed)
        if 0 <= index < len(COLOR_INDEX):
            return f"#{COLOR_INDEX[index][-6:]}".upper()

    if value.type == "theme" and value.theme is not None and theme_colors:
        theme_color = theme_colors.get(int(value.theme))
        if theme_color is not None:
            return apply_excel_tint(theme_color, float(value.tint or 0))

    return None


def cell_value_to_text(value: Any, number_format: str = "General") -> str:
    """Convert a worksheet cell value to display text."""
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, float):
        if number_format in {"", "General"}:
            if value.is_integer():
                return str(int(value))
            return format(value, ".12g")

        if "%" in number_format:
            decimal_match = re.search(r"\.([0#]+)", number_format)
            decimals = len(decimal_match.group(1)) if decimal_match else 0
            return f"{value * 100:.{decimals}f}%"

        decimal_match = re.search(r"\.([0#]+)", number_format)
        if decimal_match:
            decimals = len(decimal_match.group(1))
            grouping = "," if "," in number_format else ""
            return f"{value:{grouping}.{decimals}f}"
        if value.is_integer():
            return str(int(value))

    return str(value)


def measure_text(
    draw: Any, text: str, font: Any, stroke_width: int = 0
) -> tuple[int, int]:
    """Measure text in pixels."""
    if text == "":
        return 0, 0
    bbox = draw.textbbox((0, 0), text, font=font, stroke_width=stroke_width)
    return bbox[2] - bbox[0], bbox[3] - bbox[1]


def wrap_line(
    draw: Any, line: str, font: Any, max_width: int, stroke_width: int = 0
) -> list[str]:
    """Wrap a single text line to fit the target width."""
    if line == "":
        return [""]

    wrapped_lines = []
    current = ""
    for char in line:
        candidate = current + char
        text_width, _ = measure_text(draw, candidate, font, stroke_width)
        if current and text_width > max_width:
            wrapped_lines.append(current)
            current = char
            continue
        current = candidate

    if current:
        wrapped_lines.append(current)
    return wrapped_lines


def wrap_text(
    draw: Any, text: str, font: Any, max_width: int, stroke_width: int = 0
) -> list[str]:
    """Wrap multiline text to fit a cell."""
    lines = []
    for raw_line in text.splitlines() or [""]:
        lines.extend(wrap_line(draw, raw_line, font, max_width, stroke_width))
    return lines


def get_span_size(start: int, end: int, sizes: list[int]) -> int:
    """Return total pixel size for a one-based inclusive span."""
    return sum(sizes[start - 1 : end])


def cumulative_offsets(sizes: list[int]) -> list[int]:
    """Return leading offsets for a size list."""
    offsets = [0]
    for size in sizes:
        offsets.append(offsets[-1] + size)
    return offsets


class WorksheetImageRenderer:
    """Render openpyxl worksheets as PNG images."""

    def __init__(self, options: ImageRenderOptions) -> None:
        """Create a worksheet renderer."""
        self.options = options

    def render(self, worksheet: Any, output_path: Path) -> None:
        """Render one worksheet to a PNG image."""
        ensure_office_dependencies()
        max_row, max_col = self.get_used_bounds(worksheet)
        row_heights, col_widths = self.build_axes(worksheet, max_row, max_col)
        merged_lookup = self.build_merged_lookup(worksheet)
        group_ranges = self.build_theme_group_ranges(worksheet, max_row, max_col)
        self.add_first_column_visual_merges(merged_lookup, group_ranges)
        row_bands = self.build_row_bands(max_row, group_ranges)
        self.expand_rows_for_wrapped_text(
            worksheet, row_heights, col_widths, merged_lookup
        )
        row_offsets = cumulative_offsets(row_heights)
        col_offsets = cumulative_offsets(col_widths)

        image = Image.new(
            "RGB",
            (max(1, col_offsets[-1]), max(1, row_offsets[-1])),
            self.options.background,
        )
        draw = ImageDraw.Draw(image)
        self.draw_cells(
            draw,
            worksheet,
            row_heights,
            col_widths,
            row_offsets,
            col_offsets,
            merged_lookup,
            row_bands,
            max_row,
            max_col,
        )
        image.save(output_path)

    def scale_pixels(self, pixels: int) -> int:
        """Scale a logical pixel measurement for the output image."""
        return max(1, int(round(pixels * self.options.scale)))

    def get_scaled_padding_x(self) -> int:
        """Return horizontal cell padding at the output scale."""
        return self.scale_pixels(CELL_PADDING_X)

    def get_scaled_padding_y(self) -> int:
        """Return vertical cell padding at the output scale."""
        return self.scale_pixels(CELL_PADDING_Y)

    def get_scaled_line_spacing(self) -> int:
        """Return line spacing at the output scale."""
        return self.scale_pixels(LINE_SPACING)

    def get_used_bounds(self, worksheet: Any) -> tuple[int, int]:
        """Return worksheet bounds that should be rendered."""
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        for merged_range in worksheet.merged_cells.ranges:
            min_col, min_row, range_max_col, range_max_row = range_boundaries(
                str(merged_range)
            )
            max_col = max(max_col, min_col, range_max_col)
            max_row = max(max_row, min_row, range_max_row)
        return max_row, max_col

    def build_axes(
        self, worksheet: Any, max_row: int, max_col: int
    ) -> tuple[list[int], list[int]]:
        """Build visible row heights and column widths."""
        row_heights = []
        col_widths = []

        default_row_height = worksheet.sheet_format.defaultRowHeight
        if default_row_height is None:
            default_row_height = DEFAULT_ROW_HEIGHT_POINTS
        default_column_width = worksheet.sheet_format.defaultColWidth
        if default_column_width is None:
            default_column_width = DEFAULT_COLUMN_WIDTH

        for row_index in range(1, max_row + 1):
            dimension = worksheet.row_dimensions.get(row_index)
            row_height = (
                dimension.height
                if dimension is not None and dimension.height is not None
                else default_row_height
            )
            row_heights.append(
                0
                if dimension is not None and dimension.hidden
                else self.scale_pixels(points_to_pixels(row_height))
            )

        for col_index in range(1, max_col + 1):
            dimension = self.get_column_dimension(worksheet, col_index)
            column_width = (
                dimension.width
                if dimension is not None and dimension.width is not None
                else default_column_width
            )
            col_widths.append(
                0
                if dimension is not None and dimension.hidden
                else self.scale_pixels(column_width_to_pixels(column_width))
            )

        return row_heights, col_widths

    def get_column_dimension(self, worksheet: Any, col_index: int) -> Any:
        """Return an explicit column dimension, including grouped ranges."""
        column_letter = get_column_letter(col_index)
        dimension = worksheet.column_dimensions.get(column_letter)
        if dimension is not None:
            return dimension

        for candidate in worksheet.column_dimensions.values():
            first_column = getattr(candidate, "min", None)
            last_column = getattr(candidate, "max", None)
            if (
                first_column is not None
                and last_column is not None
                and first_column <= col_index <= last_column
            ):
                return candidate
        return None

    def build_merged_lookup(
        self, worksheet: Any
    ) -> dict[tuple[int, int], tuple[int, int, int, int]]:
        """Return a lookup from every merged coordinate to its range bounds."""
        merged_lookup = {}
        for merged_range in worksheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            for row_index in range(min_row, max_row + 1):
                for col_index in range(min_col, max_col + 1):
                    merged_lookup[(row_index, col_index)] = (
                        min_row,
                        min_col,
                        max_row,
                        max_col,
                    )
        return merged_lookup

    def find_theme_header_rows(self, worksheet: Any, max_row: int) -> list[int]:
        """Find rows where the first column starts a themed table."""
        if not self.options.use_theme_bands:
            return []

        header_rows = []
        for row_index in range(1, max_row + 1):
            cell = worksheet.cell(row_index, 1)
            text = cell_value_to_text(cell.value, cell.number_format).strip()
            if text == self.options.theme_header_text:
                header_rows.append(row_index)
        return header_rows

    def row_has_content_after_first_column(
        self, worksheet: Any, row_index: int, max_col: int
    ) -> bool:
        """Return whether a row has content outside the first column."""
        for col_index in range(2, max_col + 1):
            cell = worksheet.cell(row_index, col_index)
            if cell_value_to_text(cell.value, cell.number_format).strip():
                return True
        return False

    def next_theme_section_end(
        self, header_rows: list[int], header_index: int, max_row: int
    ) -> int:
        """Return the last row before the next themed section."""
        if header_index + 1 >= len(header_rows):
            return max_row
        return header_rows[header_index + 1] - 1

    def build_theme_group_ranges(
        self, worksheet: Any, max_row: int, max_col: int
    ) -> list[tuple[int, int]]:
        """Build first-column visual groups for themed tables."""
        header_rows = self.find_theme_header_rows(worksheet, max_row)
        group_ranges = []

        for header_index, header_row in enumerate(header_rows):
            section_end = self.next_theme_section_end(
                header_rows, header_index, max_row
            )
            group_ranges.extend(
                self.collect_theme_groups_in_section(
                    worksheet, header_row + 1, section_end, max_col
                )
            )

        return group_ranges

    def collect_theme_groups_in_section(
        self,
        worksheet: Any,
        start_row: int,
        end_row: int,
        max_col: int,
    ) -> list[tuple[int, int]]:
        """Build first-column groups inside one themed section."""
        group_ranges = []
        row_index = start_row

        while row_index <= end_row:
            cell = worksheet.cell(row_index, 1)
            text = cell_value_to_text(cell.value, cell.number_format).strip()
            has_data = self.row_has_content_after_first_column(
                worksheet, row_index, max_col
            )
            if not text or not has_data:
                row_index += 1
                continue

            group_start, group_end = self.collect_single_theme_group(
                worksheet,
                row_index,
                end_row,
                max_col,
                text,
            )
            group_ranges.append((group_start, group_end))
            row_index = group_end + 1

        return group_ranges

    def collect_single_theme_group(
        self,
        worksheet: Any,
        start_row: int,
        end_row: int,
        max_col: int,
        text: str,
    ) -> tuple[int, int]:
        """Find the final row for one first-column visual group."""
        row_index = start_row + 1
        group_end = start_row

        while row_index <= end_row:
            cell = worksheet.cell(row_index, 1)
            next_text = cell_value_to_text(cell.value, cell.number_format).strip()
            has_data = self.row_has_content_after_first_column(
                worksheet, row_index, max_col
            )
            if not has_data or (next_text and next_text != text):
                break
            group_end = row_index
            row_index += 1

        return start_row, group_end

    def add_first_column_visual_merges(
        self,
        merged_lookup: dict[tuple[int, int], tuple[int, int, int, int]],
        group_ranges: list[tuple[int, int]],
    ) -> None:
        """Add visual first-column merges for themed groups."""
        for start_row, end_row in group_ranges:
            if start_row == end_row:
                continue
            for row_index in range(start_row, end_row + 1):
                merged_lookup[(row_index, 1)] = (start_row, 1, end_row, 1)

    def build_row_bands(
        self, max_row: int, group_ranges: list[tuple[int, int]]
    ) -> dict[int, str]:
        """Build alternating row band colors."""
        row_bands = {}
        if not self.options.use_theme_bands:
            return row_bands

        for group_index, (start_row, end_row) in enumerate(group_ranges):
            band_color = DEFAULT_BAND_COLORS[group_index % len(DEFAULT_BAND_COLORS)]
            for row_index in range(start_row, end_row + 1):
                if row_index <= max_row:
                    row_bands[row_index] = band_color
        return row_bands

    def get_cell_font(self, cell: Any) -> Any:
        """Return a font sized from the cell style."""
        font_size = float(cell.font.sz or DEFAULT_FONT_SIZE)
        size = max(
            1,
            int(round(font_size * DPI_SCALE * self.options.scale)),
        )
        return load_render_font(
            size=size,
            font_name=cell.font.name,
            bold=bool(cell.font.b),
        )

    def expand_rows_for_wrapped_text(
        self,
        worksheet: Any,
        row_heights: list[int],
        col_widths: list[int],
        merged_lookup: dict[tuple[int, int], tuple[int, int, int, int]],
    ) -> None:
        """Increase row heights when wrapped text needs more room."""
        probe = Image.new("RGB", (1, 1), self.options.background)
        draw = ImageDraw.Draw(probe)

        for row in worksheet.iter_rows():
            for cell in row:
                if MergedCell is not None and isinstance(cell, MergedCell):
                    continue
                self.expand_cell_row_if_needed(
                    cell, draw, row_heights, col_widths, merged_lookup
                )

    def expand_cell_row_if_needed(
        self,
        cell: Any,
        draw: Any,
        row_heights: list[int],
        col_widths: list[int],
        merged_lookup: dict[tuple[int, int], tuple[int, int, int, int]],
    ) -> None:
        """Increase one row height when a cell wraps beyond current room."""
        text = cell_value_to_text(cell.value, cell.number_format)
        if not text:
            return

        if not cell.alignment.wrap_text:
            return

        row_index = cell.row
        col_index = cell.column
        min_row, min_col, max_row, max_col = merged_lookup.get(
            (row_index, col_index),
            (row_index, col_index, row_index, col_index),
        )
        if row_index != min_row or col_index != min_col or min_row != max_row:
            return

        font = self.get_cell_font(cell)
        cell_width = get_span_size(min_col, max_col, col_widths)
        padding_x = self.get_scaled_padding_x()
        available_width = max(self.scale_pixels(12), cell_width - padding_x * 2)
        lines = wrap_text(draw, text, font, available_width)
        _, line_height = measure_text(draw, "Ag", font)
        line_spacing = self.get_scaled_line_spacing()
        padding_y = self.get_scaled_padding_y()
        required_height = max(
            self.scale_pixels(MIN_RENDER_ROW_HEIGHT),
            len(lines) * (line_height + line_spacing) + padding_y * 2,
        )
        row_heights[row_index - 1] = max(row_heights[row_index - 1], required_height)

    def draw_cells(
        self,
        draw: Any,
        worksheet: Any,
        row_heights: list[int],
        col_widths: list[int],
        row_offsets: list[int],
        col_offsets: list[int],
        merged_lookup: dict[tuple[int, int], tuple[int, int, int, int]],
        row_bands: dict[int, str],
        max_row: int,
        max_col: int,
    ) -> None:
        """Draw worksheet cells to an image canvas."""
        drawn_ranges = set()
        for row_index in range(1, max_row + 1):
            if row_heights[row_index - 1] == 0:
                continue
            for col_index in range(1, max_col + 1):
                if col_widths[col_index - 1] == 0:
                    continue
                self.draw_cell_range(
                    draw,
                    worksheet,
                    row_offsets,
                    col_offsets,
                    merged_lookup,
                    row_bands,
                    drawn_ranges,
                    row_index,
                    col_index,
                )

    def draw_cell_range(
        self,
        draw: Any,
        worksheet: Any,
        row_offsets: list[int],
        col_offsets: list[int],
        merged_lookup: dict[tuple[int, int], tuple[int, int, int, int]],
        row_bands: dict[int, str],
        drawn_ranges: set[tuple[int, int, int, int]],
        row_index: int,
        col_index: int,
    ) -> None:
        """Draw one visible cell or merged range."""
        min_row, min_col, range_max_row, range_max_col = merged_lookup.get(
            (row_index, col_index),
            (row_index, col_index, row_index, col_index),
        )
        range_key = (min_row, min_col, range_max_row, range_max_col)
        if range_key in drawn_ranges:
            return
        drawn_ranges.add(range_key)

        cell = worksheet.cell(min_row, min_col)
        rectangle = (
            col_offsets[min_col - 1],
            row_offsets[min_row - 1],
            col_offsets[range_max_col],
            row_offsets[range_max_row],
        )
        band_color = row_bands.get(min_row)
        draw.rectangle(rectangle, fill=self.get_render_fill_color(cell, band_color))
        self.draw_border(draw, rectangle, cell)
        self.draw_cell_text(draw, cell, rectangle)

    def get_fill_color(self, cell: Any) -> str:
        """Return a cell fill color."""
        fill = cell.fill
        if fill is None or fill.fill_type is None:
            return self.options.background
        return (
            get_color(fill.fgColor, self.options.theme_colors)
            or self.options.background
        )

    def get_render_fill_color(self, cell: Any, band_color: str | None) -> str:
        """Return the cell fill color, using band color for plain cells."""
        if not self.options.preserve_fills and cell.row > 1:
            return band_color or self.options.background
        fill_color = self.get_fill_color(cell).lower()
        if band_color is not None and cell.fill.fill_type is None:
            return band_color
        return fill_color

    def draw_border(
        self, draw: Any, rectangle: tuple[int, int, int, int], cell: Any
    ) -> None:
        """Draw cell borders."""
        x1, y1, x2, y2 = rectangle
        sides = (
            (cell.border.left, (x1, y1, x1, y2)),
            (cell.border.right, (x2, y1, x2, y2)),
            (cell.border.top, (x1, y1, x2, y1)),
            (cell.border.bottom, (x1, y2, x2, y2)),
        )

        for side, line in sides:
            emphasized = side is not None and side.style in {
                "medium",
                "thick",
                "double",
            }
            color = (
                get_color(side.color, self.options.theme_colors)
                if emphasized
                else self.options.grid_color
            )
            logical_width = 2 if emphasized else 1
            width = self.scale_pixels(logical_width)
            draw.line(line, fill=color or self.options.grid_color, width=width)

    def draw_cell_text(
        self, draw: Any, cell: Any, rectangle: tuple[int, int, int, int]
    ) -> None:
        """Draw a cell value."""
        text = cell_value_to_text(cell.value, cell.number_format)
        if text == "":
            return

        x1, y1, x2, y2 = rectangle
        font = self.get_cell_font(cell)
        width = x2 - x1
        height = y2 - y1
        padding_x = self.get_scaled_padding_x()
        available_width = max(self.scale_pixels(12), width - padding_x * 2)
        lines = (
            wrap_text(draw, text, font, available_width)
            if cell.alignment.wrap_text
            else text.splitlines() or [""]
        )
        _, line_height = measure_text(draw, "Ag", font)
        line_spacing = self.get_scaled_line_spacing()
        text_height = len(lines) * (line_height + line_spacing) - line_spacing
        text_y = self.get_text_y(cell, y1, y2, height, text_height)

        color = (
            get_color(cell.font.color, self.options.theme_colors)
            or self.options.text_color
        )
        for line in lines:
            text_x = self.get_text_x(
                cell, line, draw, font, x1, x2, width
            )
            draw.text(
                (text_x, text_y),
                line,
                fill=color,
                font=font,
            )
            text_y += line_height + line_spacing

    def get_text_y(
        self, cell: Any, y1: int, y2: int, height: int, text_height: int
    ) -> int:
        """Return the y coordinate for a cell text block."""
        padding_y = self.get_scaled_padding_y()
        vertical = cell.alignment.vertical or "center"
        if vertical == "top":
            return y1 + padding_y
        if vertical == "bottom":
            return y2 - text_height - padding_y
        return y1 + max(padding_y, (height - text_height) // 2)

    def get_text_x(
        self,
        cell: Any,
        line: str,
        draw: Any,
        font: Any,
        x1: int,
        x2: int,
        width: int,
    ) -> int:
        """Return the x coordinate for one cell text line."""
        padding_x = self.get_scaled_padding_x()
        horizontal = cell.alignment.horizontal or "left"
        text_width, _ = measure_text(draw, line, font)
        if horizontal == "center":
            return x1 + max(padding_x, (width - text_width) // 2)
        if horizontal == "right":
            return x2 - text_width - padding_x
        return x1 + padding_x


class ExcelImageConverter:
    """Convert Excel workbooks to image files."""

    def convert(self, request: ImageConversionRequest) -> ImageConversionResult:
        """Export worksheet images from one workbook."""
        ensure_office_dependencies()
        workbook = load_workbook(request.input_path, data_only=True)
        request.output_dir.mkdir(parents=True, exist_ok=True)
        options = ImageRenderOptions(
            scale=request.scale,
            background=DEFAULT_BACKGROUND,
            grid_color=DEFAULT_GRID_COLOR,
            text_color=DEFAULT_TEXT_COLOR,
            theme_header_text=request.theme_header_text,
            use_theme_bands=request.use_theme_bands,
            preserve_fills=request.preserve_fills,
            theme_colors=parse_theme_colors(workbook.loaded_theme),
        )
        renderer = WorksheetImageRenderer(options)
        worksheets = self.select_worksheets(workbook, request.selected_sheets)
        output_paths = []

        for index, worksheet in enumerate(worksheets, start=1):
            filename = f"{index:02d}_{sanitize_filename(worksheet.title)}.png"
            output_path = request.output_dir / filename
            debug("Rendering worksheet", worksheet.title, "to", output_path)
            renderer.render(worksheet, output_path)
            output_paths.append(output_path)

        return ImageConversionResult(output_paths=tuple(output_paths))

    def select_worksheets(
        self, workbook: Any, selected_sheets: tuple[str, ...]
    ) -> list[Any]:
        """Return worksheets requested by name, or all worksheets when omitted."""
        if not selected_sheets:
            return list(workbook.worksheets)

        missing_sheets = [
            sheet for sheet in selected_sheets if sheet not in workbook.sheetnames
        ]
        if missing_sheets:
            missing_text = ", ".join(missing_sheets)
            raise ValueError(f"Sheet not found: {missing_text}")

        return [workbook[sheet] for sheet in selected_sheets]


class PlaceholderImageConverter:
    """Placeholder converter for Office namespaces without an implementation yet."""

    def __init__(self, namespace: str) -> None:
        """Create a placeholder converter."""
        self.namespace = namespace

    def convert(self, request: ImageConversionRequest) -> ImageConversionResult:
        """Report that this namespace is reserved for future work."""
        raise NotImplementedError(
            f"{self.namespace} convert-to-img is not implemented yet"
        )


def build_image_request(args: argparse.Namespace) -> ImageConversionRequest:
    """Build a shared image conversion request from CLI arguments."""
    return ImageConversionRequest(
        input_path=Path(args.input),
        output_dir=Path(args.output_dir),
        scale=args.scale,
        selected_sheets=tuple(args.sheet or ()),
        use_theme_bands=bool(getattr(args, "theme_bands", False)),
        preserve_fills=bool(getattr(args, "preserve_fills", False)),
        theme_header_text=args.theme_header,
    )


def run_convert_to_img(args: argparse.Namespace, converter: ImageConverter) -> int:
    """Run a shared convert-to-img command."""
    request = build_image_request(args)
    result = converter.convert(request)
    CLIStyle.write(
        f"Exported {len(result.output_paths)} image files:", CLIStyle.COLORS["TITLE"]
    )
    for output_path in result.output_paths:
        CLIStyle.write(f"  {output_path}", CLIStyle.COLORS["CONTENT"])
    return 0


def run_placeholder(args: argparse.Namespace) -> int:
    """Run a reserved placeholder command."""
    namespace = getattr(args, "namespace", "office")
    function = getattr(args, "function", "function")
    CLIStyle.write(
        f"{namespace} {function} is reserved and not implemented yet.",
        CLIStyle.COLORS["WARNING"],
    )
    return 2


def add_common_image_arguments(parser: argparse.ArgumentParser) -> None:
    """Add shared image conversion arguments to a sub-command parser."""
    parser.add_argument(
        "input",
        metavar=CLIStyle.color("INPUT", CLIStyle.COLORS["CONTENT"]),
        help=CLIStyle.color("Input Office document path.", CLIStyle.COLORS["CONTENT"]),
    )
    parser.add_argument(
        "-o",
        "--output-dir",
        default="office_images",
        metavar=CLIStyle.color("DIR", CLIStyle.COLORS["CONTENT"]),
        help=CLIStyle.color(
            "Directory for exported image files.", CLIStyle.COLORS["CONTENT"]
        ),
    )
    parser.add_argument(
        "--scale",
        type=int,
        default=2,
        choices=[1, 2, 3],
        metavar=CLIStyle.color("N", CLIStyle.COLORS["CONTENT"]),
        help=CLIStyle.color("Output image scale.", CLIStyle.COLORS["CONTENT"]),
    )


def add_excel_image_arguments(parser: argparse.ArgumentParser) -> None:
    """Add Excel-specific image conversion arguments."""
    add_common_image_arguments(parser)
    parser.add_argument(
        "--sheet",
        action="append",
        metavar=CLIStyle.color("NAME", CLIStyle.COLORS["CONTENT"]),
        help=CLIStyle.color(
            "Worksheet name to export. Repeat to export multiple sheets.",
            CLIStyle.COLORS["CONTENT"],
        ),
    )
    theme_group = parser.add_mutually_exclusive_group()
    theme_group.add_argument(
        "--theme-bands",
        dest="theme_bands",
        action="store_true",
        help=CLIStyle.color(
            "Enable optional first-column themed visual grouping.",
            CLIStyle.COLORS["CONTENT"],
        ),
    )
    theme_group.add_argument(
        "--no-theme-bands",
        dest="theme_bands",
        action="store_false",
        help=CLIStyle.color(
            "Keep only native worksheet merges and fills.",
            CLIStyle.COLORS["CONTENT"],
        ),
    )
    parser.set_defaults(theme_bands=False)
    fill_group = parser.add_mutually_exclusive_group()
    fill_group.add_argument(
        "--preserve-fills",
        dest="preserve_fills",
        action="store_true",
        help=CLIStyle.color(
            "Preserve source worksheet cell fills.", CLIStyle.COLORS["CONTENT"]
        ),
    )
    fill_group.add_argument(
        "--clean-fills",
        dest="preserve_fills",
        action="store_false",
        help=CLIStyle.color(
            "Use a neutral body background for a cleaner image.",
            CLIStyle.COLORS["CONTENT"],
        ),
    )
    parser.set_defaults(preserve_fills=False)
    parser.add_argument(
        "--theme-header",
        default=DEFAULT_THEME_HEADER_TEXT,
        metavar=CLIStyle.color("TEXT", CLIStyle.COLORS["CONTENT"]),
        help=CLIStyle.color(
            "First-column header text that starts themed grouping.",
            CLIStyle.COLORS["CONTENT"],
        ),
    )


def build_main_epilog(script_name: str) -> str:
    """Build top-level examples and notes."""
    return create_example_text(
        script_name,
        [
            (
                "Export all Excel worksheets as PNG files",
                "excel convert-to-img workbook.xlsx -o sheet_images",
            ),
            (
                "Export one worksheet at a higher scale",
                "excel convert-to-img workbook.xlsx --sheet Summary --scale 3",
            ),
            ("Show reserved Word namespace", "word placeholder"),
            ("Show reserved Slides namespace", "slides placeholder"),
        ],
        [
            "Office namespaces use: namespace function --args.",
            "Word and slides commands are currently placeholders.",
        ],
    )


def build_namespace_epilog(script_name: str, namespace: str) -> str:
    """Build examples for an Office namespace parser."""
    return create_example_text(
        script_name,
        [
            (f"Show {namespace} placeholder", f"{namespace} placeholder"),
        ],
        [
            f"{namespace} commands are reserved for future Office functions.",
        ],
    )


def build_excel_epilog(script_name: str) -> str:
    """Build examples for the Excel namespace parser."""
    return create_example_text(
        script_name,
        [
            (
                "Export all worksheets",
                "excel convert-to-img workbook.xlsx -o sheet_images",
            ),
            (
                "Export selected worksheets",
                "excel convert-to-img workbook.xlsx --sheet Summary --sheet Data",
            ),
        ],
        [
            "The converter uses openpyxl cell data and styles, not a native Excel renderer.",
        ],
    )


def build_excel_convert_epilog(script_name: str) -> str:
    """Build examples for the Excel image conversion parser."""
    return create_example_text(
        script_name,
        [
            ("Export all worksheets", "excel convert-to-img workbook.xlsx"),
            (
                "Export selected worksheets",
                "excel convert-to-img workbook.xlsx -o images --sheet Summary",
            ),
            (
                "Enable optional visual grouping",
                "excel convert-to-img workbook.xlsx --theme-bands",
            ),
            (
                "Preserve source cell fills",
                "excel convert-to-img workbook.xlsx --preserve-fills",
            ),
        ],
        [
            "Output files are PNG images named from worksheet titles.",
            "Use --scale 1, 2, or 3 to control output resolution.",
            "Body fills are normalized by default; use --preserve-fills to keep them.",
        ],
    )


def add_placeholder_parser(
    subparsers: argparse._SubParsersAction, namespace: str
) -> None:
    """Add a placeholder function parser to a namespace."""
    parser = subparsers.add_parser(
        "placeholder",
        help=CLIStyle.color(
            "Reserved placeholder function.", CLIStyle.COLORS["CONTENT"]
        ),
        description="Reserved Office function placeholder.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=build_namespace_epilog(Path(sys.argv[0]).name, namespace),
    )
    parser.set_defaults(
        handler=run_placeholder, namespace=namespace, function="placeholder"
    )


def add_convert_placeholder_parser(
    subparsers: argparse._SubParsersAction, namespace: str
) -> None:
    """Add a future convert-to-img parser for a placeholder namespace."""
    parser = subparsers.add_parser(
        "convert-to-img",
        help=CLIStyle.color(
            "Reserved image conversion function.", CLIStyle.COLORS["CONTENT"]
        ),
        description="Reserved image conversion command.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=build_namespace_epilog(Path(sys.argv[0]).name, namespace),
    )
    add_common_image_arguments(parser)
    parser.set_defaults(
        handler=lambda args: run_convert_to_img(
            args, PlaceholderImageConverter(namespace)
        ),
        namespace=namespace,
        function="convert-to-img",
        sheet=[],
        theme_bands=False,
        preserve_fills=False,
        theme_header=DEFAULT_THEME_HEADER_TEXT,
    )


def add_reserved_namespace(
    subparsers: argparse._SubParsersAction, namespace: str
) -> None:
    """Add a reserved Office namespace with placeholder functions."""
    parser = subparsers.add_parser(
        namespace,
        help=CLIStyle.color(
            f"{namespace} tools namespace.", CLIStyle.COLORS["CONTENT"]
        ),
        description=f"{namespace} Office tools.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=build_namespace_epilog(Path(sys.argv[0]).name, namespace),
    )
    function_parsers = parser.add_subparsers(
        dest="function",
        metavar=CLIStyle.color("function", CLIStyle.COLORS["CONTENT"]),
        help=CLIStyle.color("Function to run.", CLIStyle.COLORS["CONTENT"]),
    )
    add_placeholder_parser(function_parsers, namespace)
    add_convert_placeholder_parser(function_parsers, namespace)
    parser.set_defaults(
        handler=run_placeholder, namespace=namespace, function="function"
    )


def add_excel_namespace(subparsers: argparse._SubParsersAction) -> None:
    """Add the Excel namespace and functions."""
    script_name = Path(sys.argv[0]).name
    parser = subparsers.add_parser(
        "excel",
        help=CLIStyle.color("Excel tools namespace.", CLIStyle.COLORS["CONTENT"]),
        description="Excel Office tools.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=build_excel_epilog(script_name),
    )
    function_parsers = parser.add_subparsers(
        dest="function",
        metavar=CLIStyle.color("function", CLIStyle.COLORS["CONTENT"]),
        help=CLIStyle.color("Function to run.", CLIStyle.COLORS["CONTENT"]),
    )
    convert_parser = function_parsers.add_parser(
        "convert-to-img",
        aliases=["cti"],
        help=CLIStyle.color(
            "Export worksheets as PNG files.", CLIStyle.COLORS["CONTENT"]
        ),
        description="Export Excel worksheets as PNG images.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=build_excel_convert_epilog(script_name),
    )
    add_excel_image_arguments(convert_parser)
    convert_parser.set_defaults(
        handler=lambda args: run_convert_to_img(args, ExcelImageConverter()),
        namespace="excel",
        function="convert-to-img",
    )
    add_placeholder_parser(function_parsers, "excel")
    parser.set_defaults(handler=run_placeholder, namespace="excel", function="function")


def build_parser() -> argparse.ArgumentParser:
    """Build the top-level CLI parser."""
    script_name = Path(sys.argv[0]).name
    parser = ColoredArgumentParser(
        description="Office document helper toolkit.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=build_main_epilog(script_name),
    )
    parser.add_argument(
        "--log",
        action="store_true",
        help=CLIStyle.color("Enable debug output.", CLIStyle.COLORS["CONTENT"]),
    )
    namespace_parsers = parser.add_subparsers(
        dest="namespace",
        metavar=CLIStyle.color("namespace", CLIStyle.COLORS["CONTENT"]),
        help=CLIStyle.color("Office namespace to use.", CLIStyle.COLORS["CONTENT"]),
    )
    add_reserved_namespace(namespace_parsers, "word")
    add_excel_namespace(namespace_parsers)
    add_reserved_namespace(namespace_parsers, "slides")
    return parser


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    """Parse command line arguments."""
    parser = build_parser()
    args = parser.parse_args(argv)
    if not hasattr(args, "handler"):
        parser.print_help()
        raise SystemExit(2)
    return args


def main(argv: list[str] | None = None) -> int:
    """Main program logic."""
    global DEBUG_MODE
    try:
        args = parse_args(argv)
        DEBUG_MODE = bool(args.log)
        return args.handler(args)
    except FileNotFoundError as exc:
        CLIStyle.write(
            f"Error: file not found: {exc.filename}",
            CLIStyle.COLORS["ERROR"],
            error=True,
        )
        return 1
    except NotImplementedError as exc:
        CLIStyle.write(f"Error: {exc}", CLIStyle.COLORS["WARNING"], error=True)
        return 2
    except Exception as exc:
        if DEBUG_MODE:
            traceback.print_exc()
        CLIStyle.write(f"Error: {exc}", CLIStyle.COLORS["ERROR"], error=True)
        return 1


if __name__ == "__main__":
    sys.exit(main())
